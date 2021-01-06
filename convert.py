## Rutina para transformar archivos Excel en formato xlsx a texto plano csv delimitado por punto y coma(;)
## considera todas las hojas del archivo excel a procesar
## Sergio Ruiz - 04-01-2021
## XLSX a CSV

## importando las librerias necesarias
## pip install openpyxl
import sys, os
import shutil
import time
import errno
import openpyxl
import dateutil
from dateutil import parser
from openpyxl import load_workbook
from datetime import date
from datetime import datetime
import datetime

print("-----------------------> PROCESO INICIADO -----> "+time.strftime("%d/%m/%y %H:%M:%S"))

## Pidiendo la ruta donde está el archivo
print("Ingrese la ruta donde se encuentran los archivos: \n")
ejemplo_dir = input()

## Pidiendo el delimitador
print("Ingrese el delimitador de columnas: ")
delimitador = input()

## Si se quiere determinar una ruta específica
# ejemplo_dir = 'C:\\Users\\sergio.ruiz.LA.000\\Documents\\STORBOX-IM\\DMS\\SUPPORT\\CS055_BANCOCHILE\\FAQ\\_transformar_excel_csv\\'

## Seteando las rutas donde quedaran los archivos csv y excel procesados
directorio_arch_procesados = ejemplo_dir+'\\procesados'
excel_revisados = ejemplo_dir+'\\excel_revisados'

## Creando carpeta de archivos procesados si no existe
if not os.path.isdir(directorio_arch_procesados):
	os.mkdir(directorio_arch_procesados)
	print("		#### Carpeta CSV procesados se ha creado")

## Creando carpeta de archivos procesados si no existe
if not os.path.isdir(excel_revisados):
	os.mkdir(excel_revisados)
	print("		#### Carpeta Excel revisados se ha creado")

## Guardando en un arreglo los archivos encontrados que cumplan las condiciones
with os.scandir(ejemplo_dir) as archivos:
    archivos = [fichero.name for fichero in archivos if fichero.is_file() and (fichero.name.endswith('.xlsx') or fichero.name.endswith('.XLSX'))]

## Validando si existen archivos para procesar
if(len(archivos)>0):
	## Procesando cada uno de los archivos encontrados
	for x in archivos:
		filename = x

		## Abriendo el archivo xlsx
		xlsx = openpyxl.load_workbook(filename)

		##Muestra las hojas del archivo excel
		for sheetname in xlsx.sheetnames:
			sheet = xlsx[sheetname]

			## obteniendo los datos de cada hoja del archivo
			data = sheet.rows
			
			## creando archivo csv con el nombre de la hoja
			csv = open(directorio_arch_procesados+"/"+sheetname+".csv", "w+")

			for row in data:
				l = list(row)
				
				## recorre por cantidad de columnas
				for i in range(len(l)):
					## Cuando llega al final de la fila, hace un salto de linea
					if i == len(l) - 1:
						csv.write(str(l[i].value)+ '\n')
					## si no, va colocandole el separador
					else:
						# para los campos con formato fecha, se transforma en formato dd-mm-yyyy
						if (str(l[i].data_type)) == "d":
							csv.write(dateutil.parser.parse(str(l[i].value)).strftime('%d-%m-%Y') + delimitador)
						else:
							csv.write(str(l[i].value) + delimitador)
					csv.write('')

			## cerrando archivo csv
			csv.close()
		# Mueve el archivo a la carpeta de excel revisados
		print("		#### Moviendo archivo"+filename+" a la carpeta de revisados")
		shutil.move(filename, excel_revisados)
else:
	print("(WARNING!) No se encontraron archivos para procesar")

print("-----------------------> PROCESO FINALIZADO -----> "+time.strftime("%d/%m/%y %H:%M:%S"))
print("Presione cualquier tecla para salir")
gracias = input()